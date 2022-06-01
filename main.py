import mintapi
from openpyxl import load_workbook
import datetime
import time


def read_file(file_name):
    with open(file_name) as f:
        lines = f.readlines()
    for x in range(len(lines)):
        lines[x] = lines[x].strip()
    return lines


def create_mint_instance():
    creds = read_file('creds.txt')

    return mintapi.Mint(
        creds[0],  # Email used to log in to Mint
        creds[1],  # Your password used to log in to mint
        mfa_method=None,  # See MFA Methods section
        mfa_input_callback=None,  # see MFA Methods section
        mfa_token=None,  # see MFA Methods section
        intuit_account=None,  # account name when multiple accounts are registered with this email.
        headless=False,  # Whether the chromedriver should work without opening a
        session_path=None,  # Directory that the Chrome persistent session will be written/read from.
        imap_account=None,  # account name used to log in to your IMAP server
        imap_password=None,  # account password used to log in to your IMAP server
        imap_server=None,  # IMAP server host name
        imap_folder='INBOX',  # IMAP folder that receives MFA email
        wait_for_sync=False,  # do not wait for accounts to sync
        wait_for_sync_timeout=300,  # number of seconds to wait for sync
        use_chromedriver_on_path=False,  # True will use a system provided chromedriver binary that
        driver=None  # pre-configured driver. If None, Mint will initialize the WebDriver.
    )


def format_data(account_data):
    accounts = read_file('accounts.txt')
    account_dict = {}
    # This preserves the order of the accounts as desired by the textfile 'accounts.txt'
    for account in accounts:
        account_dict[account] = 0.00
    for account in account_data:
        if account['name'] in accounts and account['isActive']:
            account_dict[account['name']] = account['value']
    return account_dict


def write_info_to_excel(data_dict):
    work_book = load_workbook('Balance.xlsx')
    work_sheet = work_book.active
    next_row = work_sheet.max_row + 1
    next_col = 1

    work_sheet.cell(row=next_row, column=next_col, value=datetime.datetime.now())
    apply_formatting(work_sheet, [next_col], next_row, 'm/d/y')

    for value in data_dict.values():
        next_col += 1
        new_value = float(value.__str__().translate(str.maketrans('', '', '$,')))
        work_sheet.cell(row=next_row, column=next_col, value=new_value)
        apply_formatting(work_sheet, [next_col], next_row, '$#,##0.00')

    try:
        work_book.save('Balance.xlsx')
    except PermissionError as e:
        print(f"{e.strerror}: It appears the Excel Workbook is open. Please Close before running the script.")

    time.sleep(10)
    work_book.close()


def apply_formatting(work_sheet, cols, row, num_format):
    for col in cols:
        work_sheet[f"{num_to_column_letters(col)}{row}"].number_format = num_format


def num_to_column_letters(num):
    letters = ''
    while num:
        mod = (num - 1) % 26
        letters += chr(mod + 65)
        num = (num - 1) // 26
    return ''.join(reversed(letters))


if __name__ == '__main__':
    mint = create_mint_instance()
    acct_dict = format_data(mint.get_account_data())
    write_info_to_excel(acct_dict)
    mint.close()
    print("Success")
